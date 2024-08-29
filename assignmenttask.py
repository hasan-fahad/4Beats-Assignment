from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import datetime

# Setup WebDriver (Assuming Chrome)
driver = webdriver.Chrome()

# Load the Excel workbook
workbook = load_workbook('4BeatsQ1.xlsx')

# Get the current day of the week
current_day = datetime.datetime.now().strftime("%A")

# Load the correct sheet based on the current day
sheet = workbook[current_day]

# Iterate over the rows in the sheet
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    keyword = row[0].value
    
    # Skip if keyword is None
    if keyword is None:
        continue
    
    driver.get('https://www.google.com')
    search_box = driver.find_element(By.NAME, 'q')
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.RETURN)

    # Wait for suggestions to load
    suggestions = driver.find_elements(By.CSS_SELECTOR, 'li.sbct')

    # Get longest and shortest suggestion
    suggestions_text = [s.text for s in suggestions if s.text]
    longest_suggestion = max(suggestions_text, key=len)
    shortest_suggestion = min(suggestions_text, key=len)

    # Write the longest and shortest options back to the Excel sheet
    row[1].value = longest_suggestion
    row[2].value = shortest_suggestion

# Save the updated Excel file
workbook.save('4BeatsQ1.xlsx')

# Close the WebDriver
driver.quit()
